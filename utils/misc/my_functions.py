from openpyxl import load_workbook


def excel(excel_file: str):
    """
    :param excel_file takes a string value:
    :return list:
    """
    print("excel_file:", excel_file)
    workbook = load_workbook(excel_file)
    workbook = workbook["Жадвал"]

    sheet = workbook

    result = ""
    i = 9
    while True:
        kontur = sheet.cell(row=i, column=2)
        fosfor = sheet.cell(row=i, column=7)
        kaliy = sheet.cell(row=i, column=10)
        gumus = sheet.cell(row=i, column=12)
        if fosfor.value == None:
            break
        result += "{} - {} | {} | {} |\n".format(kontur.value, fosfor.value,
                                                 kaliy.value, gumus.value)
        i += 1
    print(result)
    return result


